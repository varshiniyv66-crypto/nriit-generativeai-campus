"""
NRIIT FACULTY DATA IMPORT
Reads faculty data from Excel files and generates SQL
"""
from openpyxl import load_workbook
import re
from datetime import datetime
import os

DATA_FOLDER = r"C:\Users\CSE_DEAN\Desktop\nriit-generativeai-campus\realupdated data of nriit"
OUTPUT_FILE = r"C:\Users\CSE_DEAN\Desktop\nriit-generativeai-campus\import_nriit_faculty.sql"

all_faculty = []

def clean_text(text):
    if text is None:
        return None
    text = str(text)
    text = re.sub(r'_x[0-9a-fA-F]{4}_', '', text)
    text = text.replace('\x00', '').replace('"', '').strip()
    return text if text else None

def extract_title(name):
    """Extract title from name"""
    name = name.upper()
    if name.startswith('DR.') or name.startswith('DR '):
        return 'Dr.', name[3:].strip()
    elif name.startswith('MR.') or name.startswith('MR '):
        return 'Mr.', name[3:].strip()
    elif name.startswith('MRS.') or name.startswith('MRS '):
        return 'Mrs.', name[4:].strip()
    elif name.startswith('MS.') or name.startswith('MS '):
        return 'Ms.', name[3:].strip()
    return '', name

def process_cse_faculty():
    """Process CSE prequalifier faculty file"""
    filepath = os.path.join(DATA_FOLDER, "cse prequalifier  faculty data 2025.xlsx")
    faculty = []
    
    try:
        wb = load_workbook(filepath, data_only=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            dept_code = 'CSE' if sheet_name == 'CSE' else ('IT' if sheet_name == 'IT' else 'ECE')
            
            # Get headers
            headers = []
            for cell in ws[1]:
                headers.append(clean_text(cell.value) if cell.value else '')
            
            # Find column indices
            name_col = -1
            degree_col = -1
            designation_col = -1
            specialization_col = -1
            pan_col = -1
            
            for i, h in enumerate(headers):
                if h:
                    h_upper = h.upper()
                    if 'NAME' in h_upper and name_col == -1:
                        name_col = i
                    elif 'DEGREE' in h_upper and degree_col == -1:
                        degree_col = i
                    elif 'DESIGNATION' in h_upper and designation_col == -1:
                        designation_col = i
                    elif 'SPECIALIZATION' in h_upper and specialization_col == -1:
                        specialization_col = i
                    elif 'PAN' in h_upper and pan_col == -1:
                        pan_col = i
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = clean_text(row[name_col]) if name_col >= 0 and name_col < len(row) else None
                if not name:
                    continue
                
                title, name_clean = extract_title(name)
                name_parts = name_clean.title().split()
                first_name = name_parts[0] if name_parts else 'Faculty'
                last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
                
                degree = clean_text(row[degree_col]) if degree_col >= 0 and degree_col < len(row) else 'M.Tech'
                designation = clean_text(row[designation_col]) if designation_col >= 0 and designation_col < len(row) else 'Assistant Professor'
                specialization = clean_text(row[specialization_col]) if specialization_col >= 0 and specialization_col < len(row) else dept_code
                pan = clean_text(row[pan_col]) if pan_col >= 0 and pan_col < len(row) else None
                
                faculty.append({
                    'first_name': first_name,
                    'last_name': last_name,
                    'title': title,
                    'dept_code': dept_code,
                    'qualification': degree,
                    'designation': designation or 'Assistant Professor',
                    'specialization': specialization,
                    'pan': pan,
                })
        
        wb.close()
    except Exception as e:
        print(f"Error: {e}")
    
    return faculty

def generate_faculty_sql(faculty):
    """Generate SQL for faculty"""
    
    sql = f"""-- ===========================================
-- NRIIT REAL FACULTY DATA IMPORT
-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}
-- Total Faculty: {len(faculty)}
-- ===========================================

"""
    
    # Group by department
    by_dept = {}
    for f in faculty:
        dept = f['dept_code']
        if dept not in by_dept:
            by_dept[dept] = []
        by_dept[dept].append(f)
    
    emp_counter = {}
    
    for dept, dept_faculty in sorted(by_dept.items()):
        sql += f"\n-- ===== {dept} Faculty: {len(dept_faculty)} =====\n"
        sql += "DO $$\nDECLARE\n    uid UUID;\nBEGIN\n"
        
        for f in dept_faculty:
            # Generate employee ID
            if dept not in emp_counter:
                emp_counter[dept] = 1
            else:
                emp_counter[dept] += 1
            
            emp_id = f"FAC{dept.replace('-', '')}{str(emp_counter[dept]).zfill(3)}"
            
            first = f['first_name'].replace("'", "''")
            last = f['last_name'].replace("'", "''")
            email = f"{emp_id.lower()}@nriit.ac.in"
            qual = f['qualification'].replace("'", "''") if f['qualification'] else 'M.Tech'
            desig = f['designation'].replace("'", "''") if f['designation'] else 'Assistant Professor'
            spec = f['specialization'].replace("'", "''") if f['specialization'] else dept
            
            sql += f"""
    -- {emp_id}: {f['title']} {first} {last}
    INSERT INTO users (email, role, is_active) VALUES ('{email}', 'faculty', true)
    ON CONFLICT (email) DO UPDATE SET last_login = NOW() RETURNING id INTO uid;
    
    INSERT INTO faculty_profiles (
        user_id, employee_id, first_name, last_name, dept_code, 
        designation, qualification, specialization, email, phone,
        date_of_joining, is_active
    ) VALUES (
        uid, '{emp_id}', '{first}', '{last}', '{dept}',
        '{desig}', '{qual}', '{spec}', '{email}', '9999999999',
        '2020-01-01', true
    ) ON CONFLICT (employee_id) DO UPDATE SET 
        first_name = EXCLUDED.first_name,
        last_name = EXCLUDED.last_name,
        designation = EXCLUDED.designation;
"""
        
        sql += "\nEND $$;\n"
    
    sql += f"\n-- SUMMARY: {len(faculty)} faculty imported\n"
    return sql

# Main
print("=" * 60)
print("NRIIT FACULTY DATA IMPORT")
print("=" * 60)

faculty = process_cse_faculty()
print(f"\n✅ Found {len(faculty)} faculty members")

# Group by department
by_dept = {}
for f in faculty:
    dept = f['dept_code']
    if dept not in by_dept:
        by_dept[dept] = 0
    by_dept[dept] += 1

for dept, count in by_dept.items():
    print(f"   {dept}: {count}")

# Generate SQL
sql = generate_faculty_sql(faculty)

with open(OUTPUT_FILE, 'w', encoding='utf-8') as file:
    file.write(sql)

print(f"\n✅ SQL saved to: {OUTPUT_FILE}")
print("🚀 Run this SQL in Supabase SQL Editor!")
