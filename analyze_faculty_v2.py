"""
Analyze Faculty Data File Structure
"""
from openpyxl import load_workbook

filepath = r"C:\Users\CSE_DEAN\Desktop\nriit-generativeai-campus\realupdated data of nriit\all l fcaualty nriit  all deapartmets 2025.xlsx"

wb = load_workbook(filepath, data_only=True)

print("=" * 60)
print("FACULTY FILE STRUCTURE ANALYSIS")
print("=" * 60)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n📋 SHEET: {sheet_name}")
    print(f"   Max Row: {ws.max_row}, Max Col: {ws.max_column}")
    
    # Print first 10 rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        # Clean values
        vals = []
        for v in row:
            if v is not None:
                v_str = str(v)
                if len(v_str) > 25:
                    v_str = v_str[:25] + "..."
                vals.append(v_str)
        if vals:
            print(f"   Row {row_idx}: {' | '.join(vals[:7])}")

wb.close()
print("\n" + "=" * 60)
