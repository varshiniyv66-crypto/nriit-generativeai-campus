"""
Analyze CSE Faculty Data File
"""
from openpyxl import load_workbook
import re

filepath = r"C:\Users\CSE_DEAN\Desktop\nriit-generativeai-campus\realupdated data of nriit\cse prequalifier  faculty data 2025.xlsx"

wb = load_workbook(filepath, data_only=True)

print("=" * 60)
print("CSE FACULTY FILE STRUCTURE")
print("=" * 60)

def clean_text(text):
    """Remove null characters and clean text"""
    if text is None:
        return None
    text = str(text)
    # Remove null chars and unicode issues
    text = re.sub(r'_x[0-9a-fA-F]{4}_', '', text)
    text = text.replace('\x00', '').strip()
    return text if text else None

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n📋 SHEET: {sheet_name}")
    print(f"   Rows: {ws.max_row}, Cols: {ws.max_column}")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        vals = []
        for v in row:
            cleaned = clean_text(v)
            if cleaned:
                if len(cleaned) > 30:
                    cleaned = cleaned[:30] + "..."
                vals.append(cleaned)
        if vals:
            print(f"   {row_idx}: {' | '.join(vals[:6])}")

wb.close()

# Also check the main faculty file with better cleaning
print("\n\n" + "=" * 60)
print("MAIN FACULTY FILE (CLEANED)")
print("=" * 60)

filepath2 = r"C:\Users\CSE_DEAN\Desktop\nriit-generativeai-campus\realupdated data of nriit\all l fcaualty nriit  all deapartmets 2025.xlsx"

wb2 = load_workbook(filepath2, data_only=True)
ws = wb2.active

for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
    vals = []
    for v in row:
        cleaned = clean_text(v)
        if cleaned:
            if len(cleaned) > 25:
                cleaned = cleaned[:25] + "..."
            vals.append(cleaned)
    if vals:
        print(f"   {row_idx}: {' | '.join(vals[:8])}")

wb2.close()
