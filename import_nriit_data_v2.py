"""
NRIIT REAL DATA IMPORT - V2
Uses openpyxl directly to read ALL cells including hidden columns
"""

import os
from openpyxl import load_workbook
from datetime import datetime

# Data folder path
DATA_FOLDER = r"C:\Users\CSE_DEAN\Desktop\nriit-generativeai-campus\realupdated data of nriit"
OUTPUT_FILE = r"C:\Users\CSE_DEAN\Desktop\nriit-generativeai-campus\import_nriit_real_data.sql"

# FINAL 10 NRIIT DEPARTMENTS (Confirmed)
DEPT_MAPPING = {
    # 1. Computer Science & Engineering
    'CSE': 'CSE', 'CSE-A': 'CSE', 'CSE -A': 'CSE', 'CSE - A': 'CSE', 'CSE-B': 'CSE', 'CSE -B': 'CSE', 'CSE - B': 'CSE', 'CSE-C': 'CSE', 'CSE -C': 'CSE', 'CSE - C': 'CSE', 'CSE-D': 'CSE', 'CSE -D': 'CSE', 'CSE - D': 'CSE', 'CSE-E': 'CSE', 'CSE -E': 'CSE',
    # 2. CSE - Data Science
    'DS': 'CSE-DS', 'DS-A': 'CSE-DS', 'DS -A': 'CSE-DS', 'DS - A': 'CSE-DS', 'DS-B': 'CSE-DS', 'DS -B': 'CSE-DS', 'DS - B': 'CSE-DS', 'DS-C': 'CSE-DS', 'DS -C': 'CSE-DS', 'DS - C': 'CSE-DS',
    # 3. Electronics & Communication Engineering
    'ECE': 'ECE', 'ECE-A': 'ECE', 'ECE -A': 'ECE', 'ECE - A': 'ECE', 'ECE-B': 'ECE', 'ECE -B': 'ECE', 'ECE - B': 'ECE', 'ECE-C': 'ECE', 'ECE -C': 'ECE', 'ECE-C': 'ECE', 'DECS': 'ECE',
    # 4. CSE - AI & ML
    'AIML': 'CSE-AI', 'AIML-A': 'CSE-AI', 'AIML-B': 'CSE-AI', 'AI': 'CSE-AI', 'AI-A': 'CSE-AI', 'AI-B': 'CSE-AI',
    # 5. Information Technology
    'IT': 'IT', 'IT-A': 'IT', 'IT-B': 'IT', 'IT -A': 'IT', 'IT - A': 'IT', 'SE': 'IT',
    # 6. Electric Vehicle Technology
    'EVT': 'EVT', 'EV': 'EVT',
    # 7. Civil Engineering
    'CE': 'CIVIL', 'CIVIL': 'CIVIL',
    # 8. BSH - Humanities & Sciences (faculty only - no students)
    'BSH': 'BSH', 'HS': 'BSH', 'H&S': 'BSH',
    # 9. MBA
    'MBA': 'MBA', 'MBA-A': 'MBA', 'MBA-B': 'MBA',
    # 10. MCA
    'MCA': 'MCA', 'MCA-A': 'MCA', 'MCA-B': 'MCA',
}

def get_dept_code(sheet_name):
    for key, value in DEPT_MAPPING.items():
        if key.upper() in sheet_name.upper().replace(' ', ''):
            return value
    return 'CSE'

def get_section(sheet_name):
    sheet_upper = sheet_name.upper()
    if '-A' in sheet_upper or ' A' == sheet_upper[-2:] or sheet_upper.endswith('-A'):
        return 'A'
    elif '-B' in sheet_upper or ' B' == sheet_upper[-2:] or sheet_upper.endswith('-B'):
        return 'B'
    elif '-C' in sheet_upper or ' C' == sheet_upper[-2:] or sheet_upper.endswith('-C'):
        return 'C'
    elif '-D' in sheet_upper or ' D' == sheet_upper[-2:] or sheet_upper.endswith('-D'):
        return 'D'
    return 'A'

def get_year_semester(filename):
    fn = filename.upper()
    if 'I B.TECH I SEM' in fn: return 1, 1
    if 'II B.TECH II SEM' in fn: return 2, 4
    if 'III B.TECH II SEM' in fn: return 3, 6
    if 'IV B.TECH II SEM' in fn: return 4, 8
    if 'MBA I SEM' in fn: return 1, 1
    if 'MBA IV SEM' in fn: return 2, 4
    if 'MCA I SEM' in fn: return 1, 1
    if 'MCA III SEM' in fn: return 2, 3
    if 'MTECH I SEM' in fn: return 1, 1
    if 'MTECH IV SEM' in fn: return 2, 4
    return 1, 1

def extract_students_from_sheet(ws, sheet_name, filename):
    """Extract student data from worksheet, reading ALL cells including hidden"""
    students = []
    dept_code = get_dept_code(sheet_name)
    section = get_section(sheet_name)
    year, semester = get_year_semester(filename)
    
    # Read all rows and find data patterns
    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=20, values_only=True):
        all_rows.append(row)
    
    # Find rows with hall ticket patterns (e.g., 24KP1A0101, 25KP1E0001)
    hall_ticket_pattern_found = False
    
    for row_idx, row in enumerate(all_rows):
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip().upper()
            
            # Check if this looks like a hall ticket number
            # Patterns: 24KP1A0101, 25KP1E0001, 22KP1D3801, etc.
            if len(cell_str) >= 10 and cell_str[:2].isdigit() and 'KP' in cell_str:
                hall_ticket = cell_str
                
                # Try to get student name from adjacent cells
                name = None
                for next_col in range(col_idx + 1, min(col_idx + 5, len(row))):
                    if row[next_col] and str(row[next_col]).strip():
                        potential_name = str(row[next_col]).strip()
                        # Check if it looks like a name (not a number, not a code)
                        if not potential_name.replace('.', '').replace(' ', '').isdigit():
                            if len(potential_name) > 3 and 'KP' not in potential_name.upper():
                                name = potential_name
                                break
                
                if name:
                    # Clean the name
                    name = name.title()
                    name_parts = name.split()
                    first_name = name_parts[0] if name_parts else 'Student'
                    last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
                    
                    # Try to get phone from other cells
                    phone = None
                    for check_col in range(len(row)):
                        if row[check_col]:
                            val = str(row[check_col])
                            if val.isdigit() and len(val) == 10 and val.startswith(('9', '8', '7', '6')):
                                phone = val
                                break
                    
                    students.append({
                        'roll_number': hall_ticket,
                        'first_name': first_name,
                        'last_name': last_name,
                        'dept_code': dept_code,
                        'section': section,
                        'year': year,
                        'semester': semester,
                        'phone': phone,
                    })
    
    return students

def process_file(filepath, filename):
    """Process an Excel file with all sheets"""
    all_students = []
    
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            students = extract_students_from_sheet(ws, sheet_name, filename)
            all_students.extend(students)
            if students:
                print(f"     Sheet '{sheet_name}': {len(students)} students")
        
        wb.close()
    except Exception as e:
        print(f"  Error: {e}")
    
    return all_students

def generate_sql(students):
    """Generate SQL for all students"""
    
    sql = f"""-- ===========================================
-- NRIIT REAL DATA IMPORT
-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}
-- Total Students: {len(students)}
-- ===========================================

-- FINAL 10 NRIIT DEPARTMENTS
INSERT INTO departments (code, name, short_name, is_active) VALUES
('CSE', 'Computer Science and Engineering', 'CSE', true),
('CSE-DS', 'CSE - Data Science', 'CSE-DS', true),
('ECE', 'Electronics and Communication Engineering', 'ECE', true),
('CSE-AI', 'CSE - Artificial Intelligence & ML', 'CSE-AI', true),
('IT', 'Information Technology', 'IT', true),
('EVT', 'Electric Vehicle Technology', 'EVT', true),
('CIVIL', 'Civil Engineering', 'CIVIL', true),
('BSH', 'Humanities and Sciences', 'BSH', true),
('MBA', 'Master of Business Administration', 'MBA', true),
('MCA', 'Master of Computer Applications', 'MCA', true)
ON CONFLICT (code) DO NOTHING;

-- Ensure batches exist
INSERT INTO academic_batches (batch_year, batch_name, is_current) VALUES
(2025, '2025-29', true), (2024, '2024-28', false), (2023, '2023-27', false), 
(2022, '2022-26', false), (2021, '2021-25', false)
ON CONFLICT DO NOTHING;

"""
    
    # Group by department
    by_dept = {}
    for s in students:
        dept = s['dept_code']
        if dept not in by_dept:
            by_dept[dept] = []
        by_dept[dept].append(s)
    
    for dept, dept_students in sorted(by_dept.items()):
        sql += f"\n-- ===== {dept}: {len(dept_students)} students =====\n"
        sql += "DO $$\nDECLARE\n    uid UUID;\n    bid UUID;\nBEGIN\n"
        
        for s in dept_students:
            roll = s['roll_number'].replace("'", "''")
            first = s['first_name'].replace("'", "''")
            last = s['last_name'].replace("'", "''")
            email = f"{roll.lower()}@nriit.ac.in"
            phone = s['phone'] if s['phone'] else '9999999999'
            section = s['section']
            sem = s['semester']
            
            batch_year = 2000 + int(roll[:2]) if roll[:2].isdigit() else 2024
            
            sql += f"""
    SELECT id INTO bid FROM academic_batches WHERE batch_year = {batch_year} LIMIT 1;
    INSERT INTO users (email, role, is_active) VALUES ('{email}', 'student', true)
    ON CONFLICT (email) DO UPDATE SET last_login = NOW() RETURNING id INTO uid;
    INSERT INTO student_profiles (user_id, roll_number, registration_number, first_name, last_name, dept_code, batch_id, current_semester, section, email, phone, admission_date, is_active)
    VALUES (uid, '{roll}', '{roll}', '{first}', '{last}', '{dept}', bid, {sem}, '{section}', '{email}', '{phone}', '{batch_year}-07-15', true)
    ON CONFLICT (roll_number) DO NOTHING;
"""
        
        sql += "\nEND $$;\n"
    
    sql += f"\n-- SUMMARY: {len(students)} students imported across {len(by_dept)} departments\n"
    return sql

# ===== MAIN =====
print("=" * 60)
print("NRIIT REAL DATA IMPORT - V2")
print("Reading ALL cells including hidden columns")
print("=" * 60)

all_students = []

files = [f for f in os.listdir(DATA_FOLDER) if f.endswith('.xlsx') and not f.startswith('~$')]

for filename in files:
    filepath = os.path.join(DATA_FOLDER, filename)
    print(f"\n📁 {filename}")
    students = process_file(filepath, filename)
    all_students.extend(students)
    print(f"   Total: {len(students)} students")

# Remove duplicates
unique = {}
for s in all_students:
    unique[s['roll_number']] = s

all_students = list(unique.values())

print(f"\n{'=' * 60}")
print(f"TOTAL UNIQUE STUDENTS: {len(all_students)}")
print("=" * 60)

# Generate SQL
sql = generate_sql(all_students)

with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    f.write(sql)

print(f"\n✅ SQL saved to: {OUTPUT_FILE}")
print("🚀 Run this SQL in Supabase SQL Editor!")
