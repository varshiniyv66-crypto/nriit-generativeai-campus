"""
COMPLETE DEPARTMENT ANALYSIS FROM NRIIT DATA FILES
Extracts and counts all departments with student counts
"""
from openpyxl import load_workbook
import os
from collections import defaultdict

DATA_FOLDER = r"C:\Users\CSE_DEAN\Desktop\nriit-generativeai-campus\realupdated data of nriit"

# Track departments and student counts
dept_counts = defaultdict(int)
dept_students = defaultdict(list)
all_sheets_found = set()

print("=" * 70)
print("NRIIT - COMPLETE DEPARTMENT EXTRACTION")
print("=" * 70)

for filename in os.listdir(DATA_FOLDER):
    if filename.endswith('.xlsx') and not filename.startswith('~$'):
        filepath = os.path.join(DATA_FOLDER, filename)
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            
            for sheet_name in wb.sheetnames:
                if sheet_name in ['Worksheet', 'Sheet3', 'Sheet1', 'Sheet2']:
                    continue  # Skip generic sheets
                    
                all_sheets_found.add(sheet_name)
                ws = wb[sheet_name]
                
                # Count students (rows with hall ticket numbers)
                student_count = 0
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=5, values_only=True):
                    for cell in row:
                        if cell:
                            cell_str = str(cell).strip().upper()
                            if len(cell_str) >= 10 and cell_str[:2].isdigit() and 'KP' in cell_str:
                                student_count += 1
                                break
                
                # Normalize sheet name to department
                sheet_upper = sheet_name.upper().strip()
                
                # Determine base department
                if 'CSE' in sheet_upper and 'DS' not in sheet_upper and 'AI' not in sheet_upper:
                    dept = 'CSE'
                elif 'DS' in sheet_upper:
                    dept = 'CSE-DS (Data Science)'
                elif 'AIML' in sheet_upper or ('AI' in sheet_upper and 'ML' in sheet_upper):
                    dept = 'CSE-AI (AI & ML)'
                elif 'IT' in sheet_upper:
                    dept = 'IT'
                elif 'ECE' in sheet_upper:
                    dept = 'ECE'
                elif 'EEE' in sheet_upper:
                    dept = 'EEE'
                elif 'MECH' in sheet_upper or 'ME' == sheet_upper:
                    dept = 'MECH'
                elif 'CE' == sheet_upper or 'CIVIL' in sheet_upper:
                    dept = 'CIVIL'
                elif 'EVT' in sheet_upper or 'EV' == sheet_upper:
                    dept = 'EVT (Electric Vehicle Tech)'
                elif 'MBA' in sheet_upper:
                    dept = 'MBA'
                elif 'MCA' in sheet_upper:
                    dept = 'MCA'
                elif 'DECS' in sheet_upper:
                    dept = 'M.Tech - DECS'
                elif 'SE' == sheet_upper:
                    dept = 'M.Tech - SE'
                else:
                    dept = f'OTHER ({sheet_name})'
                
                dept_counts[dept] += student_count
                
            wb.close()
        except Exception as e:
            print(f"Error in {filename}: {e}")

print("\n" + "=" * 70)
print("📊 DEPARTMENTS FOUND IN YOUR DATA FILES")
print("=" * 70)

# Sort by student count
sorted_depts = sorted(dept_counts.items(), key=lambda x: -x[1])

total_students = 0
dept_number = 0

print(f"\n{'No.':<4} {'Department':<35} {'Students':>10}")
print("-" * 55)

for dept, count in sorted_depts:
    if count > 0:
        dept_number += 1
        total_students += count
        print(f"{dept_number:<4} {dept:<35} {count:>10}")

print("-" * 55)
print(f"{'':4} {'TOTAL':<35} {total_students:>10}")

print("\n" + "=" * 70)
print(f"🎯 UNIQUE DEPARTMENTS WITH STUDENTS: {dept_number}")
print("=" * 70)

print("\n📋 ALL SHEET NAMES FOUND:")
for sheet in sorted(all_sheets_found):
    print(f"   • {sheet}")

print("\n" + "=" * 70)
print("SUMMARY: These are the departments extracted from your data.")
print("Please confirm which departments to include in the final system!")
print("=" * 70)
