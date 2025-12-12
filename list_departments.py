"""
List all unique departments from NRIIT data files
"""
from openpyxl import load_workbook
import os

DATA_FOLDER = r"C:\Users\CSE_DEAN\Desktop\nriit-generativeai-campus\realupdated data of nriit"

all_sheets = set()

print("=" * 60)
print("ALL DEPARTMENTS/SHEETS IN NRIIT DATA FILES")
print("=" * 60)

for filename in os.listdir(DATA_FOLDER):
    if filename.endswith('.xlsx') and not filename.startswith('~$'):
        filepath = os.path.join(DATA_FOLDER, filename)
        try:
            wb = load_workbook(filepath, read_only=True)
            print(f"\n📁 {filename}")
            for sheet in wb.sheetnames:
                all_sheets.add(sheet)
                print(f"   - {sheet}")
            wb.close()
        except Exception as e:
            print(f"   Error: {e}")

print("\n" + "=" * 60)
print(f"UNIQUE SHEETS/DEPARTMENTS: {len(all_sheets)}")
print("=" * 60)
for sheet in sorted(all_sheets):
    print(f"  • {sheet}")
